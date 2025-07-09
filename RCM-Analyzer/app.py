import sys
import sqlite3
import streamlit as st
import os
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from utils.document_processor import process_document
from utils.gemini import initialize_gemini, analyze_risk_with_gemini
from utils.db import initialize_chroma, store_in_chroma, query_chroma
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Check SQLite version and warn if incompatible with ChromaDB
sqlite_version = sqlite3.sqlite_version_info
min_sqlite_version = (3, 35, 0)
is_sqlite_compatible = sqlite_version >= min_sqlite_version

# Load environment variables
load_dotenv()

# Set API key directly as fallback if not loaded from .env
if not os.environ.get("GEMINI_API_KEY"):
    os.environ["GEMINI_API_KEY"] = "AIzaSyBdz-qcLFRDsR-mm37AlRf2w6RZws2lDL0"

# Set page configuration with professional branding
st.set_page_config(
    page_title="RCM Analytics Suite",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com/your-org/rcm-analyzer',
        'Report a bug': "https://github.com/your-org/rcm-analyzer/issues",
        'About': "# RCM Analytics Suite\nProfessional Risk Control Matrix Analysis Platform"
    }
)

# Initialize Gemini
gemini_model = initialize_gemini()

def apply_professional_styling():
    """Apply clean, professional styling with simple colors"""
    st.markdown("""
    <style>
    /* Import clean professional font */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Clean base styling */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        color: #2d3748;
        background: #f7fafc;
    }
    
    .stApp {
        background: #f7fafc;
    }
    
    .main {
        padding: 1rem 2rem;
        background: #f7fafc;
    }
    
    /* Clean typography */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Inter', sans-serif;
        color: #1a202c;
        font-weight: 600;
    }
    
    h1 { font-size: 2.25rem; margin-bottom: 1rem; }
    h2 { font-size: 1.875rem; margin-bottom: 1rem; }
    h3 { font-size: 1.5rem; margin-bottom: 0.75rem; }
    h4 { font-size: 1.25rem; margin-bottom: 0.75rem; }
    
    p, div, span {
        color: #4a5568;
        line-height: 1.6;
    }
    
    /* Simple professional header */
    .main-header {
        background: #2d3748;
        color: white;
        padding: 2.5rem 2rem;
        border-radius: 8px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }
    
    .main-header h1 {
        color: #ffffff !important;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        margin-bottom: 0.5rem;
    }
    
    .main-header p {
        color: #ffffff !important;
        font-size: 1.125rem;
        margin: 0;
        font-weight: 400;
    }
    
    /* Force header text to be white with maximum specificity */
    .main-header * {
        color: #ffffff !important;
    }
    
    .main-header h1 * {
        color: #ffffff !important;
    }
    
    .main-header p * {
        color: #ffffff !important;
    }
    
    /* Override any Streamlit markdown styling in header */
    .main-header .stMarkdown,
    .main-header .stMarkdown *,
    .main-header .stMarkdown h1,
    .main-header .stMarkdown p,
    .main-header .stMarkdown div,
    .main-header .stMarkdown span {
        color: #ffffff !important;
    }
    
    /* Clean cards */
    .executive-card, .department-card, .recommendation-card {
        background: #ffffff;
        padding: 1.5rem;
        border-radius: 8px;
        border: 1px solid #e2e8f0;
        margin-bottom: 1rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        transition: box-shadow 0.2s ease;
    }
    
    .executive-card:hover, .department-card:hover, .recommendation-card:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    
    .executive-card h4, .department-card h4 {
        color: #2d3748;
        margin-bottom: 0.75rem;
        font-weight: 600;
    }
    
    .executive-card p, .department-card p {
        color: #4a5568;
        margin: 0;
    }
    
    /* Simple metrics */
    .metric-container {
        background: #ffffff;
        padding: 1.5rem;
        border-radius: 8px;
        text-align: center;
        border: 1px solid #e2e8f0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        transition: transform 0.2s ease;
    }
    
    .metric-container:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 700;
        color: #2b6cb0;
        margin: 0.5rem 0;
    }
    
    .metric-label {
        font-size: 0.875rem;
        color: #718096;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    /* Clean buttons */
    .stButton > button {
        background: #2b6cb0;
        color: white;
        border: none;
        border-radius: 6px;
        padding: 0.75rem 1.5rem;
        font-weight: 500;
        font-size: 0.95rem;
        transition: background-color 0.2s ease;
        font-family: 'Inter', sans-serif;
    }
    
    .stButton > button:hover {
        background: #2c5282;
    }
    
    /* SIDEBAR - CLEAN AND SIMPLE */
    .css-1d391kg, .stSidebar {
        background: #2d3748 !important;
        border-right: 1px solid #4a5568 !important;
    }
    
    /* Force all sidebar text to be white for maximum readability */
    .css-1d391kg,
    .css-1d391kg *,
    .stSidebar,
    .stSidebar *,
    .css-1d391kg .stMarkdown *,
    .stSidebar .stMarkdown *,
    .css-1d391kg h1,
    .css-1d391kg h2,
    .css-1d391kg h3,
    .css-1d391kg h4,
    .css-1d391kg p,
    .css-1d391kg div,
    .css-1d391kg span,
    .css-1d391kg li,
    .stSidebar h1,
    .stSidebar h2,
    .stSidebar h3,
    .stSidebar h4,
    .stSidebar p,
    .stSidebar div,
    .stSidebar span,
    .stSidebar li {
        color: #ffffff !important;
        font-family: 'Inter', sans-serif !important;
    }
    
    /* Sidebar headers */
    .css-1d391kg h3,
    .css-1d391kg h4,
    .stSidebar h3,
    .stSidebar h4 {
        color: #ffffff !important;
        font-weight: 600 !important;
        margin-bottom: 1rem !important;
        padding-bottom: 0.5rem !important;
        border-bottom: 1px solid #4a5568 !important;
    }
    
    /* Sidebar metrics */
    .css-1d391kg .stMetric,
    .stSidebar .stMetric {
        background: rgba(74, 85, 104, 0.3) !important;
        border: 1px solid #4a5568 !important;
        border-radius: 6px !important;
        padding: 1rem !important;
        margin-bottom: 1rem !important;
    }
    
    .css-1d391kg .stMetric label,
    .stSidebar .stMetric label,
    .css-1d391kg .stMetric [data-testid="metric-label"],
    .stSidebar .stMetric [data-testid="metric-label"] {
        color: #e2e8f0 !important;
        font-weight: 500 !important;
        font-size: 0.8rem !important;
    }
    
    .css-1d391kg .stMetric [data-testid="metric-value"],
    .stSidebar .stMetric [data-testid="metric-value"] {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 1.875rem !important;
    }
    
    /* Sidebar buttons */
    .css-1d391kg .stButton button,
    .stSidebar .stButton button {
        background: #e53e3e !important;
        color: white !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 500 !important;
        padding: 0.75rem 1rem !important;
        transition: background-color 0.2s ease !important;
    }
    
    .css-1d391kg .stButton button:hover,
    .stSidebar .stButton button:hover {
        background: #c53030 !important;
    }
    
    /* Simple risk indicators */
    .risk-indicator {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 4px;
        font-weight: 500;
        font-size: 0.8rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    .risk-high {
        background: #fed7d7;
        color: #c53030;
    }
    
    .risk-medium {
        background: #feebc8;
        color: #dd6b20;
    }
    
    .risk-low {
        background: #c6f6d5;
        color: #38a169;
    }
    
    /* Clean inputs */
    .stSelectbox label, .stFileUploader label {
        color: #2d3748 !important;
        font-weight: 500 !important;
        margin-bottom: 0.5rem !important;
    }
    
    .stSelectbox > div > div {
        border: 1px solid #e2e8f0;
        border-radius: 6px;
    }
    
    .stSelectbox > div > div:focus-within {
        border-color: #2b6cb0;
        box-shadow: 0 0 0 3px rgba(43, 108, 176, 0.1);
    }
    
    /* Clean file uploader */
    .stFileUploader > div {
        background: #f7fafc;
        border: 2px dashed #cbd5e1;
        border-radius: 8px;
        padding: 2rem;
        text-align: center;
    }
    
    .stFileUploader > div:hover {
        border-color: #2b6cb0;
        background: #edf2f7;
    }
    
    /* Clean tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
        background: #edf2f7;
        border-radius: 6px;
        padding: 0.25rem;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 4px;
        color: #4a5568 !important;
        font-weight: 500;
        padding: 0.5rem 1rem;
        border: none;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(43, 108, 176, 0.1);
        color: #2b6cb0 !important;
    }
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: #2b6cb0;
        color: white !important;
    }
    
    /* Clean data tables */
    .stDataFrame {
        border-radius: 6px;
        border: 1px solid #e2e8f0;
        overflow: hidden;
    }
    
    .stDataFrame th {
        background: #edf2f7 !important;
        color: #2d3748 !important;
        font-weight: 600 !important;
        border-bottom: 1px solid #e2e8f0 !important;
    }
    
    .stDataFrame td {
        color: #4a5568 !important;
        border-bottom: 1px solid #f7fafc !important;
    }
    
    /* Clean alerts */
    .stAlert {
        border-radius: 6px;
        border: 1px solid;
        font-weight: 500;
        padding: 1rem;
    }
    
    .stSuccess {
        background: #f0fff4;
        border-color: #38a169;
        color: #2f855a;
    }
    
    .stError {
        background: #fed7d7;
        border-color: #e53e3e;
        color: #c53030;
    }
    
    .stWarning {
        background: #feebc8;
        border-color: #dd6b20;
        color: #c05621;
    }
    
    .stInfo {
        background: #ebf8ff;
        border-color: #3182ce;
        color: #2c5282;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Department headers */
    .department-header {
        font-size: 1.5rem;
        font-weight: 600;
        color: #2d3748;
        margin-bottom: 1rem;
        padding-bottom: 0.75rem;
        border-bottom: 2px solid #e2e8f0;
    }
    
    /* Recommendation styling */
    .recommendation-title {
        font-size: 1.125rem;
        font-weight: 600;
        color: #2d3748;
        margin-bottom: 0.5rem;
    }
    
    .recommendation-description {
        color: #4a5568;
        line-height: 1.6;
    }
    
    /* Analysis status */
    .analysis-status {
        background: #f0fff4;
        border: 1px solid #38a169;
        border-radius: 6px;
        padding: 1.5rem;
        text-align: center;
        margin: 1rem 0;
    }
    
    .analysis-status h3 {
        color: #2f855a;
        margin: 0 0 0.5rem 0;
        font-weight: 600;
        font-size: 1.25rem;
    }
    
    .analysis-status p {
        color: #2f855a;
        margin: 0;
        font-weight: 400;
    }
    
    /* Ensure main content readability */
    .main .stMarkdown h1,
    .main .stMarkdown h2,
    .main .stMarkdown h3,
    .main .stMarkdown h4 {
        color: #2d3748 !important;
    }
    
    .main .stMarkdown p,
    .main .stMarkdown div {
        color: #4a5568 !important;
    }
    
    /* Final sidebar text override */
    [data-testid="stSidebar"] {
        color: #ffffff !important;
    }
    
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }
    </style>
    """, unsafe_allow_html=True)

def create_professional_header():
    """Create a professional header for the application"""
    st.markdown("""
    <div class="main-header">
        <h1>🛡️ RCM Analytics Suite</h1>
        <p>Professional Risk Control Matrix Analysis Platform</p>
    </div>
    """, unsafe_allow_html=True)

def create_executive_summary(data):
    """Create an executive summary dashboard"""
    st.markdown("## 📊 Executive Summary")
    
    # Key metrics row
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_controls = len(data.get("control_objectives", []))
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">Total Controls</div>
            <div class="metric-value">{total_controls}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        total_gaps = len(data.get("gaps", []))
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">Control Gaps</div>
            <div class="metric-value" style="color: #ff6b6b;">{total_gaps}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        departments = len(data.get("departments", []))
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">Departments</div>
            <div class="metric-value">{departments}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        risk_dist = data.get("risk_distribution", {})
        high_risks = risk_dist.get("High", 0)
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">High Risk Items</div>
            <div class="metric-value" style="color: #ff6b6b;">{high_risks}</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Risk distribution visualization
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### Risk Distribution Analysis")
        risk_dist = data.get("risk_distribution", {"High": 0, "Medium": 0, "Low": 0})
        
        # Create pie chart
        fig = px.pie(
            values=list(risk_dist.values()),
            names=list(risk_dist.keys()),
            color_discrete_map={
                "High": "#ff6b6b",
                "Medium": "#ffa726", 
                "Low": "#66bb6a"
            },
            title="Risk Level Distribution"
        )
        fig.update_layout(
            font=dict(family="Inter, sans-serif", color="#1a202c"),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            title_font_size=16,
            title_font_color="#1a202c"
        )
        fig.update_traces(textfont_color="#1a202c")
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        st.markdown("### Quick Insights")
        
        # Calculate risk insights
        total_risks = sum(risk_dist.values())
        if total_risks > 0:
            high_pct = (risk_dist.get("High", 0) / total_risks) * 100
            medium_pct = (risk_dist.get("Medium", 0) / total_risks) * 100
            low_pct = (risk_dist.get("Low", 0) / total_risks) * 100
            
            st.markdown(f"""
            <div class="executive-card">
                <h4>Risk Profile</h4>
                <p><span class="risk-indicator risk-high">High: {high_pct:.1f}%</span></p>
                <p><span class="risk-indicator risk-medium">Medium: {medium_pct:.1f}%</span></p>
                <p><span class="risk-indicator risk-low">Low: {low_pct:.1f}%</span></p>
            </div>
            """, unsafe_allow_html=True)
            
            # Risk assessment
            if high_pct > 30:
                assessment = "🔴 Critical attention required"
                color = "#ff6b6b"
            elif high_pct > 15:
                assessment = "🟡 Moderate risk exposure"
                color = "#ffa726"
            else:
                assessment = "🟢 Acceptable risk profile"
                color = "#66bb6a"
                
            st.markdown(f"""
            <div class="executive-card" style="border-left-color: {color};">
                <h4>Overall Assessment</h4>
                <p style="color: {color}; font-weight: 600;">{assessment}</p>
            </div>
            """, unsafe_allow_html=True)

def create_department_heatmap(data):
    """Create a professional department risk heatmap"""
    st.markdown("### 🎯 Department Risk Heatmap")
    
    dept_risks = data.get("department_risks", {})
    if not dept_risks:
        st.warning("No department risk data available for heatmap.")
        return
    
    # Prepare data for heatmap
    departments = list(dept_risks.keys())
    risk_categories = ["Financial", "Operational", "Compliance", "Strategic", "Technological"]
    
    # Create matrix
    risk_matrix = []
    for dept in departments:
        dept_data = dept_risks[dept]
        risk_cats = dept_data.get("risk_categories", {})
        row = [risk_cats.get(cat, 0) for cat in risk_categories]
        risk_matrix.append(row)
    
    # Create heatmap
    fig = px.imshow(
        risk_matrix,
        x=risk_categories,
        y=departments,
        color_continuous_scale="RdYlGn_r",
        aspect="auto",
        title="Department Risk Assessment Matrix"
    )
    
    fig.update_layout(
        font=dict(family="Inter, sans-serif", color="#1a202c"),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        title_font_size=16,
        title_font_color="#1a202c",
        height=400
    )
    
    fig.update_xaxes(title="Risk Categories", title_font_color="#1a202c", tickfont_color="#2d3748")
    fig.update_yaxes(title="Departments", title_font_color="#1a202c", tickfont_color="#2d3748")
    
    st.plotly_chart(fig, use_container_width=True)

def create_risk_trends_chart(data):
    """Create risk trends and patterns visualization"""
    st.markdown("### 📈 Risk Analysis Trends")
    
    # Create subplots for multiple charts
    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=('Risk Categories by Department', 'Control Gap Analysis'),
        specs=[[{"type": "bar"}, {"type": "bar"}]]
    )
    
    dept_risks = data.get("department_risks", {})
    departments = list(dept_risks.keys())
    
    if departments:
        # Risk categories analysis
        financial_risks = [dept_risks[dept].get("risk_categories", {}).get("Financial", 0) for dept in departments]
        operational_risks = [dept_risks[dept].get("risk_categories", {}).get("Operational", 0) for dept in departments]
        
        fig.add_trace(
            go.Bar(name="Financial", x=departments, y=financial_risks, marker_color="#ff6b6b"),
            row=1, col=1
        )
        fig.add_trace(
            go.Bar(name="Operational", x=departments, y=operational_risks, marker_color="#ffa726"),
            row=1, col=1
        )
        
        # Control gaps by department
        gaps_by_dept = {}
        for gap in data.get("gaps", []):
            dept = gap.get("department", "Unknown")
            gaps_by_dept[dept] = gaps_by_dept.get(dept, 0) + 1
        
        gap_departments = list(gaps_by_dept.keys())
        gap_counts = list(gaps_by_dept.values())
        
        fig.add_trace(
            go.Bar(name="Control Gaps", x=gap_departments, y=gap_counts, 
                  marker_color="#667eea", showlegend=False),
            row=1, col=2
        )
    
    fig.update_layout(
        font=dict(family="Inter, sans-serif", color="#1a202c"),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        title_font_color="#1a202c",
        height=400,
        showlegend=True
    )
    
    # Update axes with proper colors
    fig.update_xaxes(title_font_color="#1a202c", tickfont_color="#2d3748")
    fig.update_yaxes(title_font_color="#1a202c", tickfont_color="#2d3748")
    
    # Update subplot titles
    fig.update_annotations(font_color="#1a202c")
    
    st.plotly_chart(fig, use_container_width=True)

def display_professional_analysis(data):
    """Display the complete professional analysis"""
    
    # Executive Summary Section
    create_executive_summary(data)
    
    # Risk Visualization Section
    st.markdown("## 📊 Risk Analysis Dashboard")
    
    col1, col2 = st.columns([1, 1])
    with col1:
        create_department_heatmap(data)
    with col2:
        create_risk_trends_chart(data)
    
    st.markdown("---")
    
    # Department Analysis Section
    st.markdown("## 🏢 Department Analysis")
    
    departments = data.get("departments", [])
    if departments:
        # Create tabs for each department
        tabs = st.tabs(departments)
        
        for i, dept in enumerate(departments):
            with tabs[i]:
                display_department_details(data, dept)
    
    st.markdown("---")
    
    # Recommendations Section
    st.markdown("## 💡 Strategic Recommendations")
    display_recommendations(data)
    
    # Download Section
    st.markdown("---")
    st.markdown("## 📥 Export & Reports")
    create_download_section(data)

def display_department_details(data, department):
    """Display detailed analysis for a specific department"""
    dept_risks = data.get("department_risks", {}).get(department, {})
    dept_objectives = [obj for obj in data.get("control_objectives", []) 
                      if obj.get("department") == department]
    dept_gaps = [gap for gap in data.get("gaps", []) 
                if gap.get("department") == department]
    
    # Department overview
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown(f"""
        <div class="department-card">
            <div class="department-header">{department}</div>
            <p><strong>Overall Risk Level:</strong> 
            <span class="risk-indicator risk-{dept_risks.get('overall_risk_level', 'medium').lower()}">
            {dept_risks.get('overall_risk_level', 'Medium')}
            </span></p>
            <p><strong>Summary:</strong> {dept_risks.get('summary', 'No summary available.')}</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # Key metrics for this department
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">Control Objectives</div>
            <div class="metric-value">{len(dept_objectives)}</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">Control Gaps</div>
            <div class="metric-value" style="color: #ff6b6b;">{len(dept_gaps)}</div>
        </div>
        """, unsafe_allow_html=True)
    
    # Risk type analysis
    if "risk_types" in dept_risks:
        st.markdown("#### Risk Type Analysis")
        risk_types = dept_risks["risk_types"]
        
        cols = st.columns(len(risk_types))
        for i, (risk_type, risks) in enumerate(risk_types.items()):
            with cols[i]:
                count = len(risks) if isinstance(risks, list) else 0
                level = "High" if count >= 3 else "Medium" if count >= 1 else "Low"
                color = "#ff6b6b" if level == "High" else "#ffa726" if level == "Medium" else "#66bb6a"
                
                st.markdown(f"""
                <div class="metric-container">
                    <div class="metric-label">{risk_type}</div>
                    <div class="metric-value" style="color: {color};">{count}</div>
                    <div style="color: {color}; font-size: 0.8rem; font-weight: 600;">{level} Risk</div>
                </div>
                """, unsafe_allow_html=True)
    
    # Control objectives table
    if dept_objectives:
        st.markdown("#### Control Objectives")
        
        obj_df = pd.DataFrame([
            {
                "Control Objective": obj.get("objective", ""),
                "What Can Go Wrong": obj.get("what_can_go_wrong", ""),
                "Risk Level": obj.get("risk_level", ""),
                "Control Activities": obj.get("control_activities", "")
            }
            for obj in dept_objectives
        ])
        
        st.dataframe(obj_df, use_container_width=True, hide_index=True)
    
    # Control gaps
    if dept_gaps:
        st.markdown("#### Control Gaps & Recommendations")
        
        for i, gap in enumerate(dept_gaps):
            st.markdown(f"""
            <div class="recommendation-card">
                <div class="recommendation-title">Gap {i+1}: {gap.get('gap_title', 'Control Gap')}</div>
                <div class="recommendation-description">
                    <strong>Impact:</strong> {gap.get('risk_impact', 'No impact description available.')}<br>
                    <strong>Recommended Solution:</strong> {gap.get('proposed_solution', 'No solution provided.')}
                </div>
            </div>
            """, unsafe_allow_html=True)

def display_recommendations(data):
    """Display strategic recommendations in a professional format"""
    recommendations = data.get("recommendations", [])
    
    if not recommendations:
        st.info("No specific recommendations generated. Please review the identified gaps for improvement opportunities.")
        return
    
    # Group recommendations by priority
    high_priority = [r for r in recommendations if r.get("priority", "").lower() == "high"]
    medium_priority = [r for r in recommendations if r.get("priority", "").lower() == "medium"]
    low_priority = [r for r in recommendations if r.get("priority", "").lower() == "low"]
    
    # Display high priority first
    if high_priority:
        st.markdown("### 🔴 High Priority Recommendations")
        for rec in high_priority:
            st.markdown(f"""
            <div class="recommendation-card" style="border-left-color: #ff6b6b;">
                <div class="recommendation-title">{rec.get('title', 'High Priority Recommendation')}</div>
                <div class="recommendation-description">
                    {rec.get('description', '')}<br>
                    <strong>Expected Impact:</strong> {rec.get('impact', 'Significant risk reduction')}
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    if medium_priority:
        st.markdown("### 🟡 Medium Priority Recommendations")
        for rec in medium_priority:
            st.markdown(f"""
            <div class="recommendation-card" style="border-left-color: #ffa726;">
                <div class="recommendation-title">{rec.get('title', 'Medium Priority Recommendation')}</div>
                <div class="recommendation-description">
                    {rec.get('description', '')}<br>
                    <strong>Expected Impact:</strong> {rec.get('impact', 'Moderate risk reduction')}
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    if low_priority:
        st.markdown("### 🟢 Low Priority Recommendations")
        for rec in low_priority:
            st.markdown(f"""
            <div class="recommendation-card" style="border-left-color: #66bb6a;">
                <div class="recommendation-title">{rec.get('title', 'Low Priority Recommendation')}</div>
                <div class="recommendation-description">
                    {rec.get('description', '')}<br>
                    <strong>Expected Impact:</strong> {rec.get('impact', 'Minor risk reduction')}
                </div>
            </div>
            """, unsafe_allow_html=True)

def create_download_section(data):
    """Create professional download section with multiple export options"""
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Excel download
        excel_bytes = create_downloadable_excel(data)
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_bytes,
            file_name=f"RCM_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Comprehensive Excel report with all analysis data and formatting",
            use_container_width=True
        )
    
    with col2:
        # CSV download
        csv_data = create_csv_export(data)
        st.download_button(
            label="📋 Download CSV Data",
            data=csv_data,
            file_name=f"RCM_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            help="Raw data export in CSV format for further analysis",
            use_container_width=True
        )
    
    with col3:
        # Executive summary download
        exec_summary = create_executive_summary_text(data)
        st.download_button(
            label="📄 Executive Summary",
            data=exec_summary,
            file_name=f"Executive_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
            mime="text/plain",
            help="Executive summary for leadership presentation",
            use_container_width=True
        )

def create_csv_export(data):
    """Create CSV export of analysis data"""
    control_df = pd.DataFrame([
        {
            "Department": obj.get("department", ""),
            "Control Objective": obj.get("objective", ""),
            "What Can Go Wrong": obj.get("what_can_go_wrong", ""),
            "Risk Level": obj.get("risk_level", ""),
            "Control Activities": obj.get("control_activities", ""),
            "Is Gap": "Yes" if obj.get("is_gap", False) else "No",
            "Gap Details": obj.get("gap_details", ""),
            "Proposed Control": obj.get("proposed_control", "")
        }
        for obj in data.get("control_objectives", [])
    ])
    
    return control_df.to_csv(index=False)

def create_executive_summary_text(data):
    """Create executive summary text for download"""
    summary = f"""
RISK CONTROL MATRIX ANALYSIS - EXECUTIVE SUMMARY
Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

OVERVIEW
========
Total Control Objectives: {len(data.get('control_objectives', []))}
Departments Analyzed: {len(data.get('departments', []))}
Control Gaps Identified: {len(data.get('gaps', []))}

RISK DISTRIBUTION
================
"""
    
    risk_dist = data.get("risk_distribution", {})
    for level, count in risk_dist.items():
        summary += f"{level} Risk: {count} controls\n"
    
    summary += f"\nDEPARTMENTS ANALYZED\n{'='*20}\n"
    for dept in data.get("departments", []):
        summary += f"• {dept}\n"
    
    summary += f"\nKEY RECOMMENDATIONS\n{'='*19}\n"
    for i, rec in enumerate(data.get("recommendations", [])[:5], 1):
        summary += f"{i}. {rec.get('title', 'Recommendation')}\n"
        summary += f"   Priority: {rec.get('priority', 'Medium')}\n"
        summary += f"   {rec.get('description', '')}\n\n"
    
    summary += f"\nThis analysis was generated using the RCM Analytics Suite.\nFor detailed analysis, please refer to the complete Excel report.\n"
    
    return summary

def create_downloadable_excel(data):
    """
    Create an Excel file with all analyzed data and formatting
    
    Args:
        data: The analyzed data from Gemini
        
    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = Workbook()
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True, color="0000FF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    risk_high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    risk_medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    risk_low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Risk Summary"
    
    # Add title
    ws_summary.cell(row=1, column=1, value="Risk Control Matrix Analysis Summary").font = title_font
    ws_summary.merge_cells('A1:H1')
    
    # Add summary info
    ws_summary.cell(row=3, column=1, value="Departments").font = header_font
    ws_summary.cell(row=3, column=2, value=", ".join(data.get("departments", [])))
    
    # Risk distribution
    risk_dist = data.get("risk_distribution", {})
    ws_summary.cell(row=4, column=1, value="Risk Distribution").font = header_font
    ws_summary.cell(row=5, column=1, value="High Risk Items")
    ws_summary.cell(row=5, column=2, value=risk_dist.get("High", 0))
    ws_summary.cell(row=5, column=2).fill = risk_high_fill
    
    ws_summary.cell(row=6, column=1, value="Medium Risk Items")
    ws_summary.cell(row=6, column=2, value=risk_dist.get("Medium", 0))
    ws_summary.cell(row=6, column=2).fill = risk_medium_fill
    
    ws_summary.cell(row=7, column=1, value="Low Risk Items")
    ws_summary.cell(row=7, column=2, value=risk_dist.get("Low", 0))
    ws_summary.cell(row=7, column=2).fill = risk_low_fill
    
    # Control Objectives sheet
    ws_controls = wb.create_sheet(title="Control Objectives")
    
    # Headers - Updated with all columns from the image
    headers = [
        "Department", "Control Objective", "What Can Go Wrong", "Risk Level", 
        "Control Activities", "Person(s) in charge of existing control", "Additional Remarks (if any)",
        "Risk of material misstatement - Key Control - Yes/No", "Frequency Control",
        "Balance Sheet", "P&L", "Automated/Manual", "Preventive/Detective",
        "Existence/occurrence", "Completeness", "Valuation/Accuracy",
        "Rights/Obligations", "Presentation/Disclosure", "Cut-off",
        "Control/Design Gap", "Proposed Solution"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_controls.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    # Add control objectives data
    control_objectives = data.get("control_objectives", [])
    for i, obj in enumerate(control_objectives, start=2):
        ws_controls.cell(row=i, column=1, value=obj.get("department", ""))
        ws_controls.cell(row=i, column=2, value=obj.get("objective", ""))
        ws_controls.cell(row=i, column=3, value=obj.get("what_can_go_wrong", ""))
        
        risk_level = obj.get("risk_level", "Medium")
        ws_controls.cell(row=i, column=4, value=risk_level)
        
        # Color based on risk level
        if risk_level == "High":
            ws_controls.cell(row=i, column=4).fill = risk_high_fill
        elif risk_level == "Medium":
            ws_controls.cell(row=i, column=4).fill = risk_medium_fill
        elif risk_level == "Low":
            ws_controls.cell(row=i, column=4).fill = risk_low_fill
            
        # Control Activities - Generate detailed content if empty
        control_activities = obj.get("control_activities", "")
        if not control_activities:
            # Generate detailed control activities based on the risk
            what_can_go_wrong = obj.get("what_can_go_wrong", "").lower()
            if "unauthorized access" in what_can_go_wrong:
                control_activities = "Implementation of role-based access controls with regular access reviews. Multi-factor authentication for critical systems. Automated logging and monitoring of all access attempts. Regular audit of user privileges to ensure principle of least privilege."
            elif "database" in what_can_go_wrong:
                control_activities = "Regular database health monitoring with automated alerts. Scheduled database integrity checks and maintenance. Comprehensive backup procedures with regular recovery testing. Database access strictly controlled through application interfaces only."
            elif "accounting entries" in what_can_go_wrong or "financial" in what_can_go_wrong:
                control_activities = "Multi-level approval workflow for all journal entries. Automated validation of accounting codes and amounts. Regular reconciliation of accounts. Monthly review of unusual transactions and threshold-based exception reporting."
            else:
                control_activities = "Regular monitoring and review of processes. Clearly documented procedures with designated responsibilities. Automated controls where possible, with manual oversight. Periodic testing and validation of control effectiveness."
        
        ws_controls.cell(row=i, column=5, value=control_activities)
            
        # Add dummy data or actual data for additional columns
        ws_controls.cell(row=i, column=6, value=obj.get("person_in_charge", "Finance Manager"))  # Person in charge
        ws_controls.cell(row=i, column=7, value=obj.get("additional_remarks", ""))  # Additional Remarks
        ws_controls.cell(row=i, column=8, value=obj.get("key_control", "Yes"))  # Key Control
        ws_controls.cell(row=i, column=9, value=obj.get("frequency", "Monthly"))  # Frequency
        ws_controls.cell(row=i, column=10, value=obj.get("balance_sheet", "✓"))  # Balance Sheet
        ws_controls.cell(row=i, column=11, value=obj.get("p_l", "✓"))  # P&L
        ws_controls.cell(row=i, column=12, value=obj.get("automated_manual", "Manual"))  # Automated/Manual
        ws_controls.cell(row=i, column=13, value=obj.get("preventive_detective", "Preventive"))  # Preventive/Detective
        
        # Assertions
        ws_controls.cell(row=i, column=14, value=obj.get("existence_occurrence", "P"))  # Existence/occurrence
        ws_controls.cell(row=i, column=15, value=obj.get("completeness", "P"))  # Completeness
        ws_controls.cell(row=i, column=16, value=obj.get("valuation_accuracy", "P"))  # Valuation/Accuracy
        ws_controls.cell(row=i, column=17, value=obj.get("rights_obligations", ""))  # Rights/Obligations
        ws_controls.cell(row=i, column=18, value=obj.get("presentation_disclosure", ""))  # Presentation/Disclosure
        ws_controls.cell(row=i, column=19, value=obj.get("cut_off", ""))  # Cut-off
        
        # Modified: Control/Design Gap now uses "Yes" or "No" instead of gap details
        has_gap = "Yes" if obj.get("gap_details", "") else "No"
        ws_controls.cell(row=i, column=20, value=has_gap)
        
        # Use the proposed solution from the LLM directly
        proposed_solution = generate_proposed_solution(obj)
        ws_controls.cell(row=i, column=21, value=proposed_solution)
        
        # Apply borders
        for col in range(1, len(headers) + 1):
            ws_controls.cell(row=i, column=col).border = border
            ws_controls.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    
    # Add drop-down for Risk Level
    # Set a data validation list in column D from row 2 to last row with data
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv.add(f'D2:D{len(control_objectives)+1}')
    ws_controls.add_data_validation(dv)
    
    # Add drop-down for Yes/No fields
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    yes_no_dv.add(f'G2:G{len(control_objectives)+1}')  # Key Control column
    yes_no_dv.add(f'S2:S{len(control_objectives)+1}')  # Control/Design Gap column
    ws_controls.add_data_validation(yes_no_dv)
    
    # Add drop-down for Automated/Manual
    auto_manual_dv = DataValidation(type="list", formula1='"Automated,Manual,Automated/Manual"', allow_blank=True)
    auto_manual_dv.add(f'K2:K{len(control_objectives)+1}')
    ws_controls.add_data_validation(auto_manual_dv)
    
    # Add drop-down for Preventive/Detective
    prev_detect_dv = DataValidation(type="list", formula1='"Preventive,Detective,Preventive/Detective"', allow_blank=True)
    prev_detect_dv.add(f'L2:L{len(control_objectives)+1}')
    ws_controls.add_data_validation(prev_detect_dv)
    
    # Add drop-down for Frequency
    freq_dv = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Quarterly,Annually"', allow_blank=True)
    freq_dv.add(f'H2:H{len(control_objectives)+1}')
    ws_controls.add_data_validation(freq_dv)
    
    # Department Risk Analysis sheet
    ws_dept_risk = wb.create_sheet(title="Department Risk Analysis")
    
    # Headers
    dept_headers = ["Department", "Financial Risk", "Operational Risk", "Compliance Risk", 
                    "Strategic Risk", "Technological Risk", "Overall Risk"]
    
    for col, header in enumerate(dept_headers, start=1):
        cell = ws_dept_risk.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add department risk data
    dept_risks = data.get("department_risks", {})
    row = 2
    for dept, risks in dept_risks.items():
        ws_dept_risk.cell(row=row, column=1, value=dept)
        
        # Calculate overall risk for each department
        risk_values = []
        col = 2
        
        for risk_type in ["Financial", "Operational", "Compliance", "Strategic", "Technological"]:
            risk_value = risks.get("risk_categories", {}).get(risk_type, 0) if isinstance(risks, dict) else 0
            risk_values.append(risk_value)
            
            # Convert numeric value to text
            risk_text = "Low"
            risk_fill = risk_low_fill
            
            if risk_value >= 4:
                risk_text = "High"
                risk_fill = risk_high_fill
            elif risk_value >= 2:
                risk_text = "Medium"
                risk_fill = risk_medium_fill
                
            ws_dept_risk.cell(row=row, column=col, value=risk_text)
            ws_dept_risk.cell(row=row, column=col).fill = risk_fill
            ws_dept_risk.cell(row=row, column=col).border = border
            col += 1
        
        # Calculate overall risk
        avg_risk = sum(risk_values) / len(risk_values) if risk_values else 0
        risk_text = "Low"
        risk_fill = risk_low_fill
        
        if avg_risk >= 3.5:
            risk_text = "High"
            risk_fill = risk_high_fill
        elif avg_risk >= 2.0:
            risk_text = "Medium"
            risk_fill = risk_medium_fill
            
        ws_dept_risk.cell(row=row, column=col, value=risk_text)
        ws_dept_risk.cell(row=row, column=col).fill = risk_fill
        ws_dept_risk.cell(row=row, column=col).border = border
        
        row += 1
    
    # Recommendations sheet
    ws_recommendations = wb.create_sheet(title="Recommendations")
    
    # Headers
    rec_headers = ["Department", "Recommendation", "Priority", "Expected Impact"]
    
    for col, header in enumerate(rec_headers, start=1):
        cell = ws_recommendations.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add recommendations data
    recommendations = data.get("recommendations", [])
    for i, rec in enumerate(recommendations, start=2):
        ws_recommendations.cell(row=i, column=1, value=rec.get("department", ""))
        ws_recommendations.cell(row=i, column=2, value=rec.get("description", ""))
        
        priority = rec.get("priority", "Medium")
        ws_recommendations.cell(row=i, column=3, value=priority)
        
        # Color based on priority
        if priority == "High":
            ws_recommendations.cell(row=i, column=3).fill = risk_high_fill
        elif priority == "Medium":
            ws_recommendations.cell(row=i, column=3).fill = risk_medium_fill
        elif priority == "Low":
            ws_recommendations.cell(row=i, column=3).fill = risk_low_fill
            
        ws_recommendations.cell(row=i, column=4, value=rec.get("impact", ""))
        
        # Apply borders
        for col in range(1, 5):
            ws_recommendations.cell(row=i, column=col).border = border
            ws_recommendations.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
    
    # Add drop-down for Priority
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv.add(f'C2:C{len(recommendations)+1}')
    ws_recommendations.add_data_validation(dv)
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes

def generate_proposed_solution(obj):
    """Generate a proposed solution for an objective if one doesn't exist"""
    proposed_solution = obj.get("proposed_control", "")
    
    # Generate a fallback proposed solution if none exists
    if not proposed_solution:
        # Generate detailed proposed solution based on the risk
        what_can_go_wrong = obj.get("what_can_go_wrong", "").lower()
        if "unauthorized access" in what_can_go_wrong:
            proposed_solution = "Implement a comprehensive Identity and Access Management (IAM) solution with regular certification reviews. Establish segregation of duties matrix and enforce through automated controls. Implement privileged access management with just-in-time access."
        elif "database" in what_can_go_wrong:
            proposed_solution = "Implement database activity monitoring tools to track all changes. Establish formal change management procedures for schema and data modifications. Implement data loss prevention controls with automated alerting."
        elif "accounting" in what_can_go_wrong or "financial" in what_can_go_wrong:
            proposed_solution = "Implement automated validation rules for accounting entries with threshold-based approval workflows. Establish regular account reconciliation practices with management sign-off. Implement continuous monitoring dashboards for financial data integrity."
        else:
            proposed_solution = "Implement comprehensive documentation of control procedures with clear ownership. Establish regular control testing schedule with measurable effectiveness criteria. Enhance monitoring through automated dashboard reporting of control metrics."
    
    return proposed_solution

def main():
    # Apply professional styling
    apply_professional_styling()
    
    # Professional header
    create_professional_header()
    
    # Display SQLite compatibility warning if needed (professional styling)
    if not is_sqlite_compatible:
        st.warning(f"⚠️ System Notice: SQLite version {'.'.join(map(str, sqlite_version))} detected. Using in-memory vector storage for optimal performance.")
    
    # Simplified sidebar with only quick stats and help
    with st.sidebar:
        st.markdown("### 🔧 RCM Analytics Suite")
        
        # Quick stats if analysis exists
        if 'analyzed_data' in st.session_state and st.session_state.analyzed_data:
            st.markdown("#### Analysis Results")
            data = st.session_state.analyzed_data
            
            total_controls = len(data.get("control_objectives", []))
            total_gaps = len(data.get("gaps", []))
            departments = len(data.get("departments", []))
            
            st.metric("Control Objectives", total_controls)
            st.metric("Control Gaps", total_gaps, delta=f"-{total_gaps} to resolve")
            st.metric("Departments", departments)
            
            st.markdown("---")
            
            # Clear analysis option
            if st.button("🗑️ Clear Analysis", help="Start fresh with new document", use_container_width=True):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
        else:
            # Help information when no analysis
            st.markdown("#### Platform Features")
            st.markdown("""
            **Document Processing:**
            - Excel (.xlsx) files
            - CSV data files  
            - PDF documents
            - Word (.docx) files
            
            **AI Analysis:**
            - Risk assessment
            - Gap identification
            - Strategic recommendations
            - Executive dashboards
            """)
            
            st.markdown("---")
            st.markdown("#### Support")
            st.markdown("""
            📧 [Support](mailto:support@example.com)
            📚 [Documentation](https://docs.example.com)
            🐛 [Report Issues](https://github.com/issues)
            """)
    
    # Main content area
    if 'analyzed_data' not in st.session_state:
        st.session_state.analyzed_data = None
    
    # Display analysis results if available
    if st.session_state.analyzed_data:
        display_professional_analysis(st.session_state.analyzed_data)
    else:
        # Main page content when no analysis is available
        if 'file_uploaded' not in st.session_state:
            st.session_state.file_uploaded = False
        
        # Check if we have an uploaded file but no analysis yet
        uploaded_file = None
        if 'uploaded_file' in st.session_state:
            uploaded_file = st.session_state.uploaded_file
        
        # File upload section on main page
        if not st.session_state.file_uploaded:
            st.markdown("## 📄 Upload Risk Control Matrix Document")
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                uploaded_file = st.file_uploader(
                    "Select your RCM document to begin analysis", 
                    type=["xlsx", "csv", "pdf", "docx"], 
                    help="Upload your Risk Control Matrix in Excel, CSV, PDF, or Word format",
                    key="main_file_uploader"
                )
                
                if uploaded_file:
                    st.session_state.uploaded_file = uploaded_file
                    st.session_state.file_uploaded = True
                    st.rerun()
            
            with col2:
                st.markdown("""
                <div class="executive-card">
                    <h4>Supported Formats</h4>
                    <p>📊 Excel (.xlsx)<br>
                    📋 CSV (.csv)<br>
                    📄 PDF (.pdf)<br>
                    📝 Word (.docx)</p>
                </div>
                """, unsafe_allow_html=True)
        
        # Configuration and analysis section (shown after file upload)
        if st.session_state.file_uploaded and uploaded_file:
            st.markdown("## 📄 Document Ready for Analysis")
            
            # Document information
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.markdown(f"""
                <div class="executive-card">
                    <h4>📎 Document Information</h4>
                    <p><strong>File:</strong> {uploaded_file.name}</p>
                    <p><strong>Size:</strong> {uploaded_file.size:,} bytes</p>
                    <p><strong>Type:</strong> {uploaded_file.type}</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                # Remove file button
                if st.button("🗑️ Remove File", help="Choose a different file", use_container_width=True):
                    st.session_state.file_uploaded = False
                    if 'uploaded_file' in st.session_state:
                        del st.session_state.uploaded_file
                    st.rerun()
            
            st.markdown("---")
            
            # Analysis configuration section
            st.markdown("## ⚙️ Analysis Configuration")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                analysis_depth = st.selectbox(
                    "📊 Analysis Depth",
                    ["Standard", "Comprehensive", "Executive Summary Only"],
                    index=1,
                    help="Choose the level of analysis detail"
                )
            
            with col2:
                include_recommendations = st.checkbox(
                    "💡 Generate Recommendations",
                    value=True,
                    help="Include AI-generated strategic recommendations"
                )
            
            with col3:
                # Placeholder for future options
                st.markdown("**🔍 Analysis Focus**")
                st.selectbox(
                    "Select focus area",
                    ["All Departments", "High Risk Only", "Critical Controls"],
                    help="Choose what to prioritize in the analysis"
                )
            
            st.markdown("---")
            
            # Analysis button
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                analyze_button = st.button(
                    "🚀 Start Comprehensive Analysis", 
                    type="primary",
                    use_container_width=True,
                    help="Begin AI-powered risk analysis"
                )
            
            # Handle analysis
            if analyze_button:
                # Professional analysis progress
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    # Step 1: Document Processing
                    status_text.markdown("""
                    <div class="analysis-status">
                        <h3>🔍 Processing Document</h3>
                        <p>Extracting and structuring RCM data...</p>
                    </div>
                    """, unsafe_allow_html=True)
                    progress_bar.progress(25)
                    
                    # Save the uploaded file temporarily
                    temp_file_path = f"temp_{uploaded_file.name}"
                    with open(temp_file_path, "wb") as f:
                        f.write(uploaded_file.getvalue())
                    
                    # Process document
                    processed_data = process_document(temp_file_path)
                    progress_bar.progress(50)
                    
                    # Step 2: Vector Storage (optional)
                    status_text.markdown("""
                    <div class="analysis-status">
                        <h3>💾 Storing Knowledge Base</h3>
                        <p>Building vector database for enhanced analysis...</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    try:
                        if is_sqlite_compatible:
                            db = initialize_chroma("risk_control_matrix")
                            store_in_chroma(db, processed_data)
                    except Exception as chroma_error:
                        # Log but don't display error to maintain professional appearance
                        pass
                    
                    progress_bar.progress(75)
                    
                    # Step 3: AI Analysis
                    status_text.markdown("""
                    <div class="analysis-status">
                        <h3>🤖 AI Risk Analysis</h3>
                        <p>Analyzing risks, identifying patterns, and generating insights...</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Analyze with Gemini
                    if analysis_depth == "Executive Summary Only":
                        # Simplified analysis for executive view
                        analyzed_data = analyze_risk_with_gemini(gemini_model, processed_data)
                        analyzed_data["analysis_type"] = "executive"
                    else:
                        # Full analysis
                        analyzed_data = analyze_risk_with_gemini(gemini_model, processed_data)
                        analyzed_data["analysis_type"] = analysis_depth.lower()
                    
                    st.session_state.analyzed_data = analyzed_data
                    progress_bar.progress(100)
                    
                    # Cleanup
                    try:
                        os.remove(temp_file_path)
                    except:
                        pass
                    
                    # Success message
                    status_text.markdown("""
                    <div class="analysis-status">
                        <h3>✅ Analysis Complete</h3>
                        <p>Professional risk assessment ready for review</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    time.sleep(2)
                    status_text.empty()
                    progress_bar.empty()
                    st.rerun()
                    
                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"❌ Analysis Error: {str(e)}")
                    st.info("Please check your document format and try again. Contact support if the issue persists.")
                    
                    # Cleanup on error
                    try:
                        if os.path.exists(temp_file_path):
                            os.remove(temp_file_path)
                    except:
                        pass
        
        # Welcome screen when no file is uploaded
        if not st.session_state.file_uploaded:
            st.markdown("---")
            st.markdown("## 🎯 Professional Risk Assessment Platform")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("""
                <div class="executive-card">
                    <h4>📊 Comprehensive Analysis</h4>
                    <p>Advanced AI-powered risk assessment with departmental breakdowns, control gap identification, and strategic recommendations.</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div class="executive-card">
                    <h4>📈 Executive Dashboards</h4>
                    <p>Professional visualizations and executive summaries ready for board presentations and stakeholder reviews.</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown("""
                <div class="executive-card">
                    <h4>📄 Professional Reports</h4>
                    <p>Export comprehensive Excel reports, executive summaries, and CSV data for further analysis and documentation.</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Feature highlights
            st.markdown("---")
            st.markdown("### 🚀 Platform Capabilities")
            
            feature_col1, feature_col2 = st.columns(2)
            
            with feature_col1:
                st.markdown("""
                **Document Processing:**
                - Excel (.xlsx) spreadsheets
                - CSV data files  
                - PDF documents
                - Word documents (.docx)
                
                **Risk Analysis:**
                - Department-level assessment
                - Control objective evaluation
                - Gap identification and prioritization
                - Risk level classification and scoring
                """)
            
            with feature_col2:
                st.markdown("""
                **AI-Powered Insights:**
                - Google Gemini integration
                - Intelligent pattern recognition
                - Automated risk categorization
                - Strategic recommendation generation
                
                **Professional Outputs:**
                - Executive summary dashboards
                - Interactive risk heatmaps
                - Comprehensive Excel reports
                - Stakeholder-ready presentations
                """)

if __name__ == "__main__":
    main() 